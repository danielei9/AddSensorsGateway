# coding: utf-8
import openpyxl
import addSensorToGateway
path = "./test.xlsx"
from datetime import datetime
import time

gateway = "dca632143f21"
print(""" Seleccione un modo de uso: \n     1. Agregar sensores mediante excel... (Configuracion del excel 1º columna componentName y 3º Deveui)\n     2. Manualamente 1by1 () \n     3. Get Servers \n     4. Get Oems \n     5. Delete from id to id\n     6. PUT from id to id\n     7. PUT from list id """)
mode = input()
print(" Introduzca la mac gateway... ejemplo b827eb495f0c")
gateway = input()
addSensorToGateway.initLog(gateway)

if(gateway == ""):
    gateway = "dca632143f21"
if(mode == "1" or mode == "2"):

    print("Introduzca Nombre Modelo Ex( BmetersBmeters SJEvoSJEvo)")
    typelogy = input()
    print("Introduzca Nombre App Ex( app app2 )")
    appName = input()
    print("Introduzca server Id Ex( 1 2 )")
    serverId = input()
    print("Introduzca sensorName  Ex( volume  or currentMeter -> SJEVO )")
    sensorName = input()
    print("Introduzca Sensor Id  Ex( S01 )")
    sensorId = input()
if(mode == "1"):
    print("Incluya el path del excel: (ejemplo: C:/Users/dbs99/Desktop/InstalacionAras.xlsx)")
    path = input()
    if(path==""):
        path = "./test.xlsx"
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    loraAddress = []
    componentName = []
    
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        componentName.append(cell_obj.value)
    #  print(cell_obj.value)

    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 3)
        loraAddress.append(cell_obj.value)
    # print(cell_obj.value)
    print(loraAddress[1])
    print(componentName[1])
    
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
   
    for x in range(1,len(loraAddress)):
        print( str(loraAddress[x]) + " " + str(componentName[x]) )
        #print("ok?")    
        #input()
                # "body" : "{\\"sensorModelName\\":\\\"""" + typelogy + """\\",\\"applicationName\\":\\\"""" + appName + """\\"\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\\"""" + sensorName + """\\":\\\"""" + sensorId + """\\",\\"serverIds\\":["""+ serverId + """]}",

        # body" : "{\\"sensorModelName\\":\\"SJEvoSJEvo\\",\\"applicationName\\":\\"app\\",\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\"volume\\":\\"S01\\"},\\"serverIds\\":[2]}",
        messageToSend = """{
"path" : "/oemsensors",
"method" : "POST",
"body" : "{\\"sensorModelName\\":\\\"""" + typelogy + """\\",\\"applicationName\\":\\\"""" + appName + """\\",\\"loraAddress\\":\\\"""" + loraAddress[x] + """\\",\\"componentName\\":\\\"""" + componentName[x] + """\\",\\"sensorNames\\":{\\\"""" + sensorName + """\\":\\\"""" + sensorId + """\\"},\\"serverIds\\":["""+ serverId + """]}",
"port" : 4999,
"timestamp" : "2019-12-08T16:00:02.2805625Z",
"requestId" : "123456790",
"authentication" :true
}"""
        addSensorToGateway.send(gateway,messageToSend)
        time.sleep(0.1)
        time.sleep(0.1)
        time.sleep(0.1)
        time.sleep(0.1)
        time.sleep(0.1)
        time.sleep(0.1)
        time.sleep(0.1)
if(mode == "2"):
    print("Lora Address DEVEUI: ")
    loraAddressStr = input()
    print("Component Name: ")
    componentNameStr = input()
    messageToSend = """{
    "path" : "/oemsensors",
    "method" : "POST",
    "body" : "{\\"sensorModelName\\":\\\"""" + typelogy + """\\",\\"applicationName\\":\\\"""" + appName + """\\",\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\\"""" + sensorName + """\\":\\\"""" + sensorId + """\\"},\\"serverIds\\":["""+ serverId + """]}",
    "port" : 4999,
    "timestamp" : "2019-12-08T16:00:02.2805625Z",
    "requestId" : "123456790",
    "authentication" :true
}"""

    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    time.sleep(2) 
    addSensorToGateway.send(gateway,messageToSend)
    print("CTRL + C To exit2")
    while True:
        c=0
        time.sleep(0.1)

if(mode == "3"):
    messageToSend = """{
        "path" : "/servers",
        "method" : "GET",
        "body" : "",
        "port" : 4999,
        "timestamp" : "2019-12-08T16:00:02.2805625Z",
        "requestId" : "123456790",
        "authentication" :true
    }"""
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    #addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    time.sleep(0.5) 
    addSensorToGateway.send(gateway,messageToSend)
    print("CTRL + C To exit2")
    while True:
        c=0
        time.sleep(0.1)

if(mode == "4"):
    messageToSend = """{ 
    "path" : "/oemsensors", 
    "method" : "GET", 
    "body" : "", 
    "port" : 4999, 
    "timestamp" : "2019-12-08T16:00:02.2805625Z", 
    "requestId" : "123456790", 
    "authentication" :true 
    }"""
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    time.sleep(0.5) 
    addSensorToGateway.send(gateway,messageToSend)
    print("CTRL + C To exit2")
    while True:
        c=0
        time.sleep(0.1)

    

if(mode == "5"):
    print(" eliminar funciona por rango")
    print("introduzca desde que id desea eliminar")
    startId = input()
    print("introduzca hasta que id desea eliminar")
    finalId = input()
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    for i in range(int(startId),int(finalId)):
            # body" : "{\\"sensorModelName\\":\\"SJEvoSJEvo\\",\\"applicationName\\":\\"app\\",\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\"volume\\":\\"S01\\"},\\"serverIds\\":[2]}",
        messageToSend = """{ 
    "path" : "/oemsensors/""" + str(i) + """\", 
    "method" : "DELETE", 
    "body" : "", 
    "port" : 4999, 
    "timestamp" : "2019-12-08T16:00:02.2805625Z", 
    "requestId" : "123456790", 
    "authentication" :true 
    }"""
        addSensorToGateway.send(gateway,messageToSend)
        
    while True:
        c=0
        time.sleep(0.1)



if(mode == "6"):
    print(" put funciona por rango")
    print("introduzca desde que id desea eliminar")
    startId = input()
    print("introduzca hasta que id desea eliminar")
    finalId = input()
    print("introduzca Nombre de la variable Ejemplo CurrentMeter")
    nameVar = input()
    print("introduzca Variable Ejemplo : S01 ")
    var = input()
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    for i in range(int(startId),int(finalId)):
            # body" : "{\\"sensorModelName\\":\\"SJEvoSJEvo\\",\\"applicationName\\":\\"app\\",\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\"volume\\":\\"S01\\"},\\"serverIds\\":[2]}",
        messageToSend = """{ 
    "path" : "/oemsensors/""" + str(i) + """\", 
    "method" : "PUT", 
    "body" : "{\"sensorNames\":{\"""" + str(nameVar) + """\":\"""" + str(var) + """\"}}",
    "port" : 4999, 
    "timestamp" : "2019-12-08T16:00:02.2805625Z", 
    "requestId" : "123456790", 
    "authentication" :true 
    }"""
        addSensorToGateway.send(gateway,messageToSend)
    
    while True:
        c=0
        time.sleep(0.1)
    
if(mode == "7"):
    print(" put funciona por lista")
    print("introduzca las id separadas por un espacio")
    idliststr = input()
    idlist = idliststr.split(" ")
    print(idlist)
    print("introduzca Nombre de la variable Ejemplo CurrentMeter")
    nameVar = input()
    print("introduzca Variable Ejemplo : S01 ")
    var = input()
    addSensorToGateway.createLoopMqttRecive(gateway + "/gateway_requests/response/#")
    for i in idlist:
            # body" : "{\\"sensorModelName\\":\\"SJEvoSJEvo\\",\\"applicationName\\":\\"app\\",\\"loraAddress\\":\\\"""" + loraAddressStr + """\\",\\"componentName\\":\\\"""" + componentNameStr + """\\",\\"sensorNames\\":{\\"volume\\":\\"S01\\"},\\"serverIds\\":[2]}",
        messageToSend = """{ 
    "path" : "/oemsensors/""" + str(i) + """\", 
    "method" : "PUT", 
    "body" : "{\"sensorNames\":{\"""" + str(nameVar) + """\":\"""" + str(var) + """\"}}",
    "port" : 4999, 
    "timestamp" : "2019-12-08T16:00:02.2805625Z", 
    "requestId" : "123456790", 
    "authentication" :true 
    }"""
        print(i)
        addSensorToGateway.send(gateway,messageToSend)
    
    while True:
        c=0
        time.sleep(0.1)
    